from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def formatar_planilha(arquivo_saida):
    """
    Aplica toda a formatação na planilha de output.
    """

    wb = load_workbook(arquivo_saida)
    ws = wb.active

    # ===============================
    # Cabeçalho
    # ===============================

    azul = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    fonte = Font(
        color="FFFFFF",
        bold=True
    )

    for celula in ws[1]:

        celula.fill = azul
        celula.font = fonte
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    # ===============================
    # Largura das colunas
    # ===============================

    for coluna in ws.columns:

        letra = get_column_letter(
            coluna[0].column
        )

        maior = 0

        for celula in coluna:

            if celula.value is not None:

                maior = max(
                    maior,
                    len(str(celula.value))
                )

        ws.column_dimensions[letra].width = min(
            maior + 4,
            60
        )

    # ===============================
    # Localizar colunas
    # ===============================

    coluna_valor = None
    coluna_data = None

    for celula in ws[1]:

        texto = str(
            celula.value
        ).strip().upper()

        if texto == "VALOR_ESTIMADO":
            coluna_valor = celula.column

        elif texto == "DATA_DISPUTA":
            coluna_data = celula.column

    # ===============================
    # Formatação do valor
    # ===============================

    if coluna_valor:

        for linha in range(
            2,
            ws.max_row + 1
        ):

            celula = ws.cell(
                row=linha,
                column=coluna_valor
            )

            if isinstance(
                celula.value,
                (int, float)
            ):

                celula.number_format = (
                    '[$R$-416] #,##0.00'
                )

    # ===============================
    # Formatação da data
    # ===============================

    if coluna_data:

        for linha in range(
            2,
            ws.max_row + 1
        ):

            celula = ws.cell(
                row=linha,
                column=coluna_data
            )

            celula.number_format = "DD/MM/YYYY"

    # ===============================
    # Alinhamento
    # ===============================

    for linha in ws.iter_rows():

        for celula in linha:

            celula.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    wb.save(arquivo_saida)